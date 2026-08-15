"""Tests op het weggeschreven Excel-bestand.

Deze tests openen het echte XLSX-bestand. Ze controleren de opmaakeisen uit de
opdracht: bevroren kop, filters, geen samengevoegde cellen, echte getallen,
echte datums, echte booleans, klikbare URLs en consistente tellingen tussen de
tabbladen.
"""

import csv
import os

import pytest
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "outputs", "vc-investeringen-volledig.xlsx")
CSV = os.path.join(ROOT, "outputs", "vc-investeringen-volledig.csv")

SHEETS = ["README", "Fondsen", "Investeringen", "Rondes", "Portefeuillebedrijven",
          "Dekking", "Bronnen", "Conflicten", "Onbekend", "Aliases"]

REQUIRED_COLUMNS = [
    "investment_id", "round_id", "fund_slug", "fund_name", "fund_name_in_source",
    "project_slug", "project_name", "project_name_in_source", "previous_project_name",
    "announcement_date", "round_date", "round_type", "investment_type", "fund_role",
    "is_lead", "round_size_usd", "valuation_usd", "valuation_type", "fund_ticket_usd",
    "currency_original", "amount_original", "country", "sector", "chain_or_ecosystem",
    "token_ticker", "token_exists", "acquisition_or_exit", "acquirer", "exit_price_usd",
    "primary_source_url", "secondary_source_url", "aggregator_source_url",
    "source_consulted_date", "verification_status", "confidence", "conflict_flag", "notes",
]


@pytest.fixture(scope="module")
def wb():
    if not os.path.exists(XLSX):
        pytest.skip("workbook ontbreekt; draai eerst build_workbook.py")
    return load_workbook(XLSX)


def test_opent_zonder_waarschuwing(wb):
    assert wb.sheetnames == SHEETS


def test_alle_verplichte_kolommen(wb):
    header = [c.value for c in wb["Investeringen"][1]]
    for col in REQUIRED_COLUMNS:
        assert col in header, col


def test_kop_bevroren_en_filter_op_iedere_datatabel(wb):
    for name in SHEETS[1:]:
        ws = wb[name]
        assert ws.freeze_panes == "A2", name
        assert ws.tables or ws.auto_filter.ref, name


def test_geen_samengevoegde_cellen(wb):
    for name in SHEETS:
        assert not wb[name].merged_cells.ranges, name


def test_bedragen_zijn_getallen_en_nooit_nul(wb):
    ws = wb["Investeringen"]
    header = [c.value for c in ws[1]]
    for field in ("round_size_usd", "valuation_usd", "fund_ticket_usd", "exit_price_usd"):
        i = header.index(field)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            v = row[i].value
            assert v is None or isinstance(v, (int, float)), (field, v)
            assert v != 0, (field, row[0].value)


def test_datums_zijn_datumcellen(wb):
    ws = wb["Investeringen"]
    header = [c.value for c in ws[1]]
    i = header.index("round_date")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        v = row[i].value
        assert v is None or hasattr(v, "year")


def test_booleans_zijn_booleans(wb):
    ws = wb["Investeringen"]
    header = [c.value for c in ws[1]]
    for field in ("is_lead", "token_exists", "conflict_flag"):
        i = header.index(field)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            v = row[i].value
            assert v is None or isinstance(v, bool), (field, v)


def test_urls_klikbaar(wb):
    ws = wb["Investeringen"]
    header = [c.value for c in ws[1]]
    i = header.index("aggregator_source_url")
    with_link = sum(1 for row in ws.iter_rows(min_row=2, max_row=ws.max_row)
                    if row[i].value and row[i].hyperlink)
    total = sum(1 for row in ws.iter_rows(min_row=2, max_row=ws.max_row) if row[i].value)
    assert with_link == total and total > 0


def test_voorwaardelijke_opmaak_op_conflicten(wb):
    """Conflicten rood, onzekere records geel — twee regels over dezelfde range.

    openpyxl bundelt regels met dezelfde range onder één sleutel, dus tel de
    regels zelf en niet het aantal ranges.
    """
    ws = wb["Investeringen"]
    rules = [r for rule_list in ws.conditional_formatting._cf_rules.values()
             for r in rule_list]
    assert len(rules) >= 2
    formulas = " ".join(f for r in rules for f in (r.formula or []))
    assert "conflict_flag" in formulas or "TRUE" in formulas
    assert "uncertain" in formulas


def test_twintig_fondsen_op_fondsen_dekking_en_aliases(wb):
    assert wb["Aliases"].max_row - 1 == 20
    assert wb["Dekking"].max_row - 1 == 20
    slugs = {wb["Fondsen"].cell(row=r, column=1).value
             for r in range(2, 22)}
    assert len(slugs) == 20


def test_tellingen_consistent_tussen_tabbladen(wb):
    inv = wb["Investeringen"]
    header = [c.value for c in inv[1]]
    ri = header.index("round_id")
    fi = header.index("fund_slug")
    round_ids = set()
    per_fund = {}
    for row in inv.iter_rows(min_row=2, max_row=inv.max_row):
        round_ids.add(row[ri].value)
        per_fund[row[fi].value] = per_fund.get(row[fi].value, 0) + 1

    rounds_sheet = {wb["Rondes"].cell(row=r, column=1).value
                    for r in range(2, wb["Rondes"].max_row + 1)}
    assert round_ids == rounds_sheet

    for r in range(2, 22):
        slug = wb["Fondsen"].cell(row=r, column=1).value
        count = wb["Fondsen"].cell(row=r, column=4).value
        assert count == per_fund.get(slug, 0), slug


def test_dekking_telt_niet_hoger_dan_investeringen(wb):
    inv = wb["Investeringen"]
    header = [c.value for c in inv[1]]
    fi, pi = header.index("fund_slug"), header.index("project_slug")
    companies = {}
    for row in inv.iter_rows(min_row=2, max_row=inv.max_row):
        companies.setdefault(row[fi].value, set()).add(row[pi].value)
    for r in range(2, 22):
        slug = wb["Dekking"].cell(row=r, column=1).value
        own = wb["Dekking"].cell(row=r, column=3).value
        assert own == len(companies.get(slug, set())), slug


PER_FUND = os.path.join(ROOT, "outputs", "per-fonds")


@pytest.fixture(scope="module")
def fund_slugs():
    import sys
    sys.path.insert(0, os.path.join(ROOT, "scripts"))
    from funds import FUNDS
    return list(FUNDS)


def test_een_bestand_per_fonds(fund_slugs):
    if not os.path.isdir(PER_FUND):
        pytest.skip("per-fonds map ontbreekt; draai eerst build_workbook.py")
    for slug in fund_slugs:
        assert os.path.exists(os.path.join(PER_FUND, "vc-investeringen-%s.xlsx" % slug)), slug
    assert len([f for f in os.listdir(PER_FUND) if f.endswith(".xlsx")]) == 20


def test_fondsbestand_bevat_alleen_eigen_fonds(fund_slugs):
    if not os.path.isdir(PER_FUND):
        pytest.skip("per-fonds map ontbreekt")
    for slug in fund_slugs:
        fwb = load_workbook(os.path.join(PER_FUND, "vc-investeringen-%s.xlsx" % slug))
        assert fwb.sheetnames == SHEETS, slug
        ws = fwb["Investeringen"]
        header = [c.value for c in ws[1]]
        i = header.index("fund_slug")
        found = {row[i].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row)}
        assert found <= {slug}, (slug, found)
        assert fwb["Aliases"].max_row - 1 == 1, slug
        assert fwb["Dekking"].max_row - 1 == 1, slug


def test_fondsbestanden_tellen_op_tot_het_overzicht(wb, fund_slugs):
    if not os.path.isdir(PER_FUND):
        pytest.skip("per-fonds map ontbreekt")
    total = 0
    for slug in fund_slugs:
        ws = load_workbook(os.path.join(
            PER_FUND, "vc-investeringen-%s.xlsx" % slug))["Investeringen"]
        total += ws.max_row - 1 if ws.max_row > 1 else 0
    assert total == wb["Investeringen"].max_row - 1


def test_fondsbestand_houdt_co_investeerders(fund_slugs):
    """Rondes toont álle investeerders, ook fondsen buiten dit bestand."""
    if not os.path.isdir(PER_FUND):
        pytest.skip("per-fonds map ontbreekt")
    ws = load_workbook(os.path.join(PER_FUND, "vc-investeringen-paradigm.xlsx"))["Rondes"]
    header = [c.value for c in ws[1]]
    i = header.index("all_investors")
    values = [row[i].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row)]
    # Minstens één ronde noemt medeinvesteerders naast Paradigm zelf.
    assert any(v and ";" in v for v in values)


def test_csv_komt_overeen_met_investeringen(wb):
    if not os.path.exists(CSV):
        pytest.skip("CSV ontbreekt")
    with open(CSV, encoding="utf-8") as fh:
        rows = list(csv.reader(fh))
    assert len(rows) - 1 == wb["Investeringen"].max_row - 1
    assert rows[0][0] == "investment_id"
    for col in REQUIRED_COLUMNS:
        assert col in rows[0], col
