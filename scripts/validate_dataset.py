#!/usr/bin/env python3
"""Validates the dataset and the workbook. Exits with code 1 on failure.

Checks the dataset itself AND the written XLSX file, so a bug in the write
layer does not go unnoticed.
"""

import json
import os
import sys
from datetime import date

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "scripts"))
from funds import FUNDS  # noqa: E402

PROCESSED = os.path.join(ROOT, "data", "processed")
XLSX = os.path.join(ROOT, "outputs", "vc-investments-full.xlsx")
CSV = os.path.join(ROOT, "outputs", "vc-investments-full.csv")
PER_FUND = os.path.join(ROOT, "outputs", "per-fund")

ALLOWED = {
    "fund_role": {"lead", "co-lead", "participant", "incubator", "unknown"},
    "investment_type": {"equity", "token", "SAFT", "public_sale", "strategic",
                        "incubated", "unknown"},
    "verification_status": {"verified_primary", "verified_two_sources",
                            "verified_aggregator_only", "single_source",
                            "conflict", "uncertain"},
    "confidence": {"high", "medium", "low"},
    "valuation_type": {"pre_money", "post_money", "FDV", "enterprise_value", "unknown", ""},
}

failures = []
notes = []


def check(condition, message, detail=""):
    if condition:
        print("  ok    %s" % message)
    else:
        print("  FAIL  %s %s" % (message, detail))
        failures.append(message)


def main():
    with open(os.path.join(PROCESSED, "dataset.json")) as fh:
        ds = json.load(fh)

    inv = ds["investments"]
    rounds = {r["round_id"] for r in ds["rounds"]}
    fund_slugs = {f["fund_slug"] for f in ds["funds"]}
    projects = {p["project_slug"] for p in ds["projects"]}

    print("Dataset")
    ids = [r["investment_id"] for r in inv]
    check(len(ids) == len(set(ids)), "investment_id is unique",
          "(%d rows, %d unique)" % (len(ids), len(set(ids))))

    pairs = [(r["fund_slug"], r["round_id"]) for r in inv]
    check(len(pairs) == len(set(pairs)), "fund_slug + round_id is unique",
          "(%d pairs, %d unique)" % (len(pairs), len(set(pairs))))

    missing_round = [r["investment_id"] for r in inv if r["round_id"] not in rounds]
    check(not missing_round, "every round_id exists in Rounds",
          "(%d missing)" % len(missing_round))

    missing_fund = sorted({r["fund_slug"] for r in inv if r["fund_slug"] not in fund_slugs})
    check(not missing_fund, "every fund_slug exists in Funds", str(missing_fund))

    empty_project = [r["investment_id"] for r in inv if not r["project_slug"]]
    check(not empty_project, "project_slug is not blank", "(%d blank)" % len(empty_project))

    missing_project = sorted({r["project_slug"] for r in inv if r["project_slug"] not in projects})
    check(not missing_project, "every project_slug exists in Portfolio Companies",
          "(%d missing)" % len(missing_project))

    no_source = [r["investment_id"] for r in inv
                 if not (r["primary_source_url"] or r["secondary_source_url"]
                         or r["aggregator_source_url"])]
    check(not no_source, "every investment row has at least one source URL",
          "(%d without a source)" % len(no_source))

    bad_lead = [r["investment_id"] for r in inv
                if r["is_lead"] != (r["fund_role"] in ("lead", "co-lead"))]
    check(not bad_lead, "is_lead matches fund_role", "(%d deviating)" % len(bad_lead))

    ticket_eq = [r["investment_id"] for r in inv
                 if r["fund_ticket_usd"] is not None
                 and r["fund_ticket_usd"] == r["round_size_usd"]]
    check(not ticket_eq, "fund_ticket_usd is not automatically equal to round_size_usd",
          "(%d equal)" % len(ticket_eq))

    negative = [r["investment_id"] for r in inv
                for f in ("round_size_usd", "valuation_usd", "fund_ticket_usd", "exit_price_usd")
                if r.get(f) is not None and r[f] <= 0]
    check(not negative, "amounts are positive when present", "(%d wrong)" % len(negative))

    val_type = [r["investment_id"] for r in inv
                if r["valuation_usd"] and not r["valuation_type"]]
    check(not val_type, "valuation_type is filled in when valuation_usd is present",
          "(%d blank)" % len(val_type))

    conflict_ids = {c["investment_id"] for c in ds["conflicts"]}
    orphan = [r["investment_id"] for r in inv
              if r["conflict_flag"] and r["investment_id"] not in conflict_ids]
    check(not orphan, "every conflict_flag has a row in Conflicts",
          "(%d without a row)" % len(orphan))

    for field, allowed in ALLOWED.items():
        bad = sorted({str(r.get(field)) for r in inv if str(r.get(field)) not in allowed})
        check(not bad, "%s only uses allowed categories" % field, str(bad[:5]))

    missing_funds = sorted(set(FUNDS) - {r["fund_slug"] for r in inv})
    if missing_funds:
        notes.append("funds without an investment row: %s" % ", ".join(missing_funds))
    check(len(fund_slugs) == len(FUNDS), "all funds in funds.py are in Funds",
          "(%d vs %d in funds.py)" % (len(fund_slugs), len(FUNDS)))

    print("\nWorkbook")
    check(os.path.exists(XLSX), "XLSX exists")
    if not os.path.exists(XLSX):
        return finish()

    wb = load_workbook(XLSX)
    expected = ["README", "Funds", "Investments", "Rounds", "Portfolio Companies",
                "Coverage", "Sources", "Conflicts", "Unknown", "Aliases"]
    check(wb.sheetnames == expected, "sheets are in the prescribed order",
          str(wb.sheetnames))

    ws = wb["Investments"]
    check(ws.freeze_panes == "A2", "first row frozen on Investments",
          str(ws.freeze_panes))
    check(len(ws.tables) == 1 or ws.auto_filter.ref, "filter present on Investments")
    check(not ws.merged_cells.ranges, "no merged cells on Investments")
    check(ws.max_row - 1 == len(inv), "row count on Investments matches",
          "(sheet %d, dataset %d)" % (ws.max_row - 1, len(inv)))

    header = [c.value for c in ws[1]]
    for required in ["investment_id", "round_id", "fund_slug", "round_size_usd",
                     "valuation_usd", "fund_ticket_usd", "primary_source_url",
                     "verification_status", "conflict_flag"]:
        check(required in header, "column %s present" % required)

    # cell types sampled across the whole table
    idx = {name: i + 1 for i, name in enumerate(header)}
    bad_amount, bad_date, bad_bool, zero_amount = 0, 0, 0, 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        a = row[idx["round_size_usd"] - 1].value
        if a is not None and not isinstance(a, (int, float)):
            bad_amount += 1
        if a == 0:
            zero_amount += 1
        d = row[idx["round_date"] - 1].value
        if d is not None and not hasattr(d, "year"):
            bad_date += 1
        b = row[idx["is_lead"] - 1].value
        if b is not None and not isinstance(b, bool):
            bad_bool += 1
    check(bad_amount == 0, "amounts are numeric cells", "(%d wrong)" % bad_amount)
    check(zero_amount == 0, "no zero as unknown round size", "(%d zeros)" % zero_amount)
    check(bad_date == 0, "dates are date cells", "(%d wrong)" % bad_date)
    check(bad_bool == 0, "booleans are real booleans", "(%d wrong)" % bad_bool)

    links = 0
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 400)):
        if row[idx["aggregator_source_url"] - 1].hyperlink:
            links += 1
    check(links > 0, "URLs are clickable", "(%d hyperlinks in the first 400 rows)" % links)

    # cross-sheet counts
    check(wb["Rounds"].max_row - 1 == len(ds["rounds"]), "row count on Rounds matches")
    check(wb["Portfolio Companies"].max_row - 1 == len(ds["projects"]),
          "row count on Portfolio Companies matches")
    check(wb["Conflicts"].max_row - 1 == len(ds["conflicts"]) or not ds["conflicts"],
          "row count on Conflicts matches")
    check(wb["Unknown"].max_row - 1 == len(ds["unknowns"]) or not ds["unknowns"],
          "row count on Unknown matches")
    n = len(FUNDS)
    check(wb["Aliases"].max_row - 1 == n, "%d alias rows" % n)
    check(wb["Coverage"].max_row - 1 == n, "%d coverage rows" % n)
    check(wb["Funds"].max_row - 1 >= n, "%d fund rows" % n)

    sum_fund_counts = sum(
        r[3].value or 0 for r in wb["Funds"].iter_rows(min_row=2, max_row=n + 1)
    )
    check(sum_fund_counts == len(inv),
          "sum of investments_in_database equals the number of investment rows",
          "(%d vs %d)" % (sum_fund_counts, len(inv)))

    check(os.path.exists(CSV), "CSV exists")
    if os.path.exists(CSV):
        with open(CSV, encoding="utf-8") as fh:
            lines = sum(1 for _ in fh)
        check(lines - 1 == len(inv), "CSV has as many rows as Investments",
              "(%d vs %d)" % (lines - 1, len(inv)))

    print("\nPer-fund workbooks")
    per_fund_expected = {}
    for r in inv:
        per_fund_expected.setdefault(r["fund_slug"], []).append(r)

    check(os.path.isdir(PER_FUND), "outputs/per-fund directory exists")
    files = sorted(f for f in os.listdir(PER_FUND) if f.endswith(".xlsx")) \
        if os.path.isdir(PER_FUND) else []
    check(len(files) == len(FUNDS), "%d fund files" % len(FUNDS), "(%d found)" % len(files))

    total_rows = 0
    for slug in FUNDS:
        path = os.path.join(PER_FUND, "vc-investments-%s.xlsx" % slug)
        if not os.path.exists(path):
            check(False, "file for %s exists" % slug)
            continue
        fwb = load_workbook(path)
        if fwb.sheetnames != expected:
            check(False, "%s: sheets in the prescribed order" % slug,
                  str(fwb.sheetnames))
            continue
        fws = fwb["Investments"]
        expected_rows = per_fund_expected.get(slug, [])
        rows_in_sheet = fws.max_row - 1 if fws.max_row > 1 else 0
        total_rows += rows_in_sheet

        fheader = [c.value for c in fws[1]]
        ok_cols = fheader == header
        ok_count = rows_in_sheet == len(expected_rows)
        fi = fheader.index("fund_slug") if "fund_slug" in fheader else None
        only_own = True
        if fi is not None:
            for row in fws.iter_rows(min_row=2, max_row=fws.max_row):
                if row[fi].value not in (None, slug):
                    only_own = False
                    break
        ok_alias = fwb["Aliases"].max_row - 1 == 1
        ok_coverage = fwb["Coverage"].max_row - 1 == 1
        ok_freeze = fws.freeze_panes == "A2"
        ok_rounds = fwb["Rounds"].max_row - 1 == len({r["round_id"] for r in expected_rows})

        problems = []
        if not ok_cols:
            problems.append("columns differ")
        if not ok_count:
            problems.append("rows %d vs %d" % (rows_in_sheet, len(expected_rows)))
        if not only_own:
            problems.append("contains another fund")
        if not ok_alias:
            problems.append("Aliases not one row")
        if not ok_coverage:
            problems.append("Coverage not one row")
        if not ok_freeze:
            problems.append("header not frozen")
        if not ok_rounds:
            problems.append("round count wrong")
        check(not problems, "%s: fund file consistent (%d rows)"
              % (slug, rows_in_sheet), "; ".join(problems))

    check(total_rows == len(inv),
          "sum of all fund files equals the overview",
          "(%d vs %d)" % (total_rows, len(inv)))

    return finish()


def finish():
    print("")
    for n in notes:
        print("  note: %s" % n)
    if failures:
        print("VALIDATION FAILED: %d check(s) failed" % len(failures))
        return 1
    print("VALIDATION PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
