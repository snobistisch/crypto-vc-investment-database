#!/usr/bin/env python3
"""Builds outputs/vc-investments-full.xlsx and the matching CSV.

The workbook is built entirely from `data/processed/dataset.json`. No value
is typed by hand in this script: every cell comes from the dataset or is a
formula over the sheets.

Formatting rules: first row frozen, autofilter on every data table, no merged
cells, amounts as real numbers, dates as real dates, booleans as TRUE/FALSE,
clickable URLs, conflicting and uncertain records get conditional formatting.
"""

import csv
import json
import os
import sys
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "scripts"))
from funds import FUNDS  # noqa: E402

PROCESSED = os.path.join(ROOT, "data", "processed")
OUTPUTS = os.path.join(ROOT, "outputs")

XLSX = os.path.join(OUTPUTS, "vc-investments-full.xlsx")
CSV = os.path.join(OUTPUTS, "vc-investments-full.csv")
PER_FUND = os.path.join(OUTPUTS, "per-fund")

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
CONFLICT_FILL = PatternFill("solid", fgColor="F8D7DA")
UNCERTAIN_FILL = PatternFill("solid", fgColor="FFF3CD")
LINK_FONT = Font(color="0563C1", underline="single", size=10)
BODY_FONT = Font(size=10)
THIN = Side(style="thin", color="D0D7DE")

USD_FMT = '#,##0;[Red]-#,##0'
DATE_FMT = "yyyy-mm-dd"

COMPLETENESS_STATEMENT = (
    "Complete within the publicly accessible and named sources as of the "
    "cutoff date. Unannounced rounds, secondary transactions, liquid market "
    "positions, and investors that press releases lump under 'others' remain "
    "structurally invisible."
)

INVESTMENT_COLUMNS = [
    ("investment_id", "txt", 18),
    ("round_id", "txt", 18),
    ("fund_slug", "txt", 20),
    ("fund_name", "txt", 22),
    ("fund_name_in_source", "txt", 26),
    ("project_slug", "txt", 26),
    ("project_name", "txt", 26),
    ("project_name_in_source", "txt", 26),
    ("previous_project_name", "txt", 22),
    ("announcement_date", "date", 16),
    ("round_date", "date", 13),
    ("round_type", "txt", 14),
    ("investment_type", "txt", 15),
    ("fund_role", "txt", 12),
    ("is_lead", "bool", 9),
    ("round_size_usd", "usd", 16),
    ("valuation_usd", "usd", 16),
    ("valuation_type", "txt", 14),
    ("fund_ticket_usd", "usd", 15),
    ("currency_original", "txt", 10),
    ("amount_original", "usd", 15),
    ("country", "txt", 12),
    ("sector", "txt", 14),
    ("chain_or_ecosystem", "txt", 18),
    ("token_ticker", "txt", 12),
    ("token_exists", "bool", 12),
    ("acquisition_or_exit", "txt", 18),
    ("acquirer", "txt", 16),
    ("exit_price_usd", "usd", 15),
    ("primary_source_url", "url", 42),
    ("secondary_source_url", "url", 30),
    ("aggregator_source_url", "url", 42),
    ("source_consulted_date", "date", 18),
    ("verification_status", "txt", 24),
    ("confidence", "txt", 11),
    ("conflict_flag", "bool", 12),
    ("co_investors", "txt", 50),
    ("notes", "txt", 50),
]


def as_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def write_table(ws, name, columns, rows, freeze="A2"):
    """Writes one data table with header, formatting, autofilter and frozen header."""
    for idx, (title, _kind, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 22

    for r_i, row in enumerate(rows, start=2):
        for c_i, (title, kind, _w) in enumerate(columns, start=1):
            value = row.get(title)
            cell = ws.cell(row=r_i, column=c_i)
            cell.font = BODY_FONT
            cell.border = Border(bottom=THIN)
            if kind == "usd":
                # An unknown amount stays blank. Zero is not an unknown value.
                if value in (None, "", 0):
                    cell.value = None
                else:
                    cell.value = value
                    cell.number_format = USD_FMT
            elif kind == "date":
                d = as_date(value) if isinstance(value, str) else value
                if d:
                    cell.value = d
                    cell.number_format = DATE_FMT
            elif kind == "bool":
                cell.value = bool(value) if value is not None else None
            elif kind == "url":
                if value:
                    cell.value = value
                    cell.hyperlink = value
                    cell.font = LINK_FONT
            elif kind == "int":
                cell.value = value if value not in (None, "") else None
                cell.number_format = "#,##0"
            else:
                cell.value = value if value not in (None, "") else None
                cell.alignment = Alignment(vertical="top", wrap_text=(_w >= 40))

    last_col = get_column_letter(len(columns))
    last_row = max(2, len(rows) + 1)
    ref = "A1:%s%d" % (last_col, last_row)
    if rows:
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)
    else:
        ws.auto_filter.ref = ref
    ws.freeze_panes = freeze
    return last_row


def sheet_readme(wb, dataset, controls, fund_slug=None):
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 110

    if fund_slug:
        title = "Crypto VC Investment Database — %s" % FUNDS[fund_slug]["fund_name"]
        scope = [
            ("Scope of this file", "Only %s. The full overview with all twenty funds is in "
                                    "vc-investments-full.xlsx." % FUNDS[fund_slug]["fund_name"]),
            ("Co-investors", "The Rounds sheet shows ALL investors in the same round, including "
                             "funds outside this file. That is deliberate: the rest of the cap "
                             "table is exactly the interesting part."),
        ]
    else:
        title = "Crypto VC Investment Database"
        scope = [
            ("Scope of this file", "All twenty funds. A separate file also exists per fund in "
                                    "outputs/per-fund/."),
        ]

    lines = [
        (title, ""),
        ("", ""),
    ] + scope + [
        ("", ""),
        ("Source cutoff date", dataset["source_consulted_date"]),
        ("Workbook generated", dataset["generated_at"]),
        ("Funds in this file", str(len(dataset["funds"]))),
        ("", ""),
        ("What this file is", "One row per combination of fund and funding round, built from "
                              "round pages, not from fund pages."),
        ("Completeness", COMPLETENESS_STATEMENT),
        ("", ""),
        ("Method", "The round universe was fetched in a single pass over every project page on "
                   "crypto-fundraising.info; the index was then inverted to fund. Fund pages "
                   "structurally show ten rows and were therefore used only as a control list."),
        ("Projects searched", "%d" % dataset["scrape"]["projects_parsed"]),
        ("Rounds in the source", "%d" % dataset["scrape"]["rounds_total_in_source"]),
        ("Investor relations in the source", "%d" % dataset["scrape"]["investor_edges_total"]),
        ("", ""),
        ("Sheets", ""),
        ("Funds", "The twenty funds, with control totals per external source."),
        ("Investments", "One row per fund-round pair. This is the main table and the CSV export."),
        ("Rounds", "One row per round. A round with five selected funds appears here once and "
                   "in Investments five times."),
        ("Portfolio Companies", "One row per company in which at least one of the twenty funds "
                                "invested."),
        ("Coverage", "Own count next to the official portfolio page, crypto-fundraising.info, "
                     "RootData and CryptoRank."),
        ("Sources", "Every source URL used, with the date consulted and its type."),
        ("Conflicts", "Fields on which sources contradict each other. Not smoothed over."),
        ("Unknown", "Records with missing fields and what was attempted."),
        ("Aliases", "Canonical name, source name and the reason names were merged."),
        ("", ""),
        ("How to read this", ""),
        ("Blank cell", "Value unknown. Never replaced with zero or an estimate."),
        ("round_size_usd", "Size of the entire round. Explicitly not the amount the fund itself "
                           "put in."),
        ("fund_ticket_usd", "Amount the fund itself put in. Almost never public; therefore "
                            "almost always blank."),
        ("valuation_usd", "Valuation of the round as stated by the source, with valuation_type "
                          "next to it. Never a current token FDV."),
        ("", ""),
        ("Colour coding", "Red: conflict_flag is TRUE. Yellow: verification status uncertain or "
                          "single_source."),
        ("", ""),
        ("Source priority", "1 official company announcement, 2 fund announcement, 3 official "
                            "document, 4 press release, 5 crypto-fundraising.info, 6 RootData, "
                            "7 CryptoRank, 8 media."),
        ("", ""),
        ("Disclaimer", "Not investment advice."),
    ]
    for i, (label, value) in enumerate(lines, start=1):
        a = ws.cell(row=i, column=1, value=label or None)
        b = ws.cell(row=i, column=2, value=value or None)
        a.font = Font(bold=True, size=10)
        b.font = BODY_FONT
        b.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.freeze_panes = "A2"
    return ws


def scope_to_fund(dataset, fund_slug):
    """Restricts the dataset to one fund, keeping the same structure as the whole.

    Rounds, companies, conflicts and unknown fields are filtered along, so the
    counts within a fund file are internally consistent. `all_investors` on
    the Rounds sheet is left unchanged: the other investors in the same round
    are exactly the interesting part and are not filtered out.
    """
    inv = [r for r in dataset["investments"] if r["fund_slug"] == fund_slug]
    round_ids = {r["round_id"] for r in inv}
    project_slugs = {r["project_slug"] for r in inv}
    return {
        "generated_at": dataset["generated_at"],
        "source_consulted_date": dataset["source_consulted_date"],
        "scrape": dataset["scrape"],
        "funds": [f for f in dataset["funds"] if f["fund_slug"] == fund_slug],
        "investments": inv,
        "rounds": [r for r in dataset["rounds"] if r["round_id"] in round_ids],
        "projects": [p for p in dataset["projects"] if p["project_slug"] in project_slugs],
        "conflicts": [c for c in dataset["conflicts"] if c["fund_slug"] == fund_slug],
        "unknowns": [u for u in dataset["unknowns"] if u["fund_slug"] == fund_slug],
        "aliases": [a for a in dataset["aliases"] if a["fund_slug"] == fund_slug],
    }


def build(dataset, controls, xlsx_path, csv_path=None, fund_slug=None):
    """Builds one workbook. With `fund_slug`, the scope is a single fund."""
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb, dataset, controls, fund_slug)

    investments = sorted(
        dataset["investments"],
        key=lambda r: (r["fund_name"], r["round_date"] or "", r["project_name"]),
    )
    rounds = sorted(dataset["rounds"], key=lambda r: (r["round_date"] or "", r["project_name"]))
    projects = sorted(dataset["projects"], key=lambda r: r["project_name"].lower())

    # --- Funds
    per_fund = {}
    for inv in investments:
        f = per_fund.setdefault(inv["fund_slug"], {"rounds": 0, "companies": set(), "leads": 0})
        f["rounds"] += 1
        f["companies"].add(inv["project_slug"])
        if inv["is_lead"]:
            f["leads"] += 1

    fund_rows = []
    for f in dataset["funds"]:
        slug = f["fund_slug"]
        stats = per_fund.get(slug, {"rounds": 0, "companies": set(), "leads": 0})
        ctrl = controls.get("funds", {}).get(slug, {})
        fund_rows.append({
            "fund_slug": slug,
            "fund_name": f["fund_name"],
            "source_slugs": f["source_slugs"],
            "investments_in_database": stats["rounds"],
            "unique_companies": len(stats["companies"]),
            "lead_rounds": stats["leads"],
            "cryptorank_investments": (ctrl.get("cryptorank") or {}).get("investments"),
            "country_per_cryptorank": (ctrl.get("cryptorank") or {}).get("country") or "",
            "official_portfolio_url": f["official_portfolio_url"],
            "aggregator_fund_url": f["aggregator_fund_url"],
            "cryptorank_url": f["cryptorank_url"],
        })
    ws = wb.create_sheet("Funds")
    write_table(ws, "Funds", [
        ("fund_slug", "txt", 20), ("fund_name", "txt", 22), ("source_slugs", "txt", 28),
        ("investments_in_database", "int", 20), ("unique_companies", "int", 17),
        ("lead_rounds", "int", 12), ("cryptorank_investments", "int", 20),
        ("country_per_cryptorank", "txt", 20),
        ("official_portfolio_url", "url", 40), ("aggregator_fund_url", "url", 40),
        ("cryptorank_url", "url", 40),
    ], fund_rows)

    # --- Investments
    ws = wb.create_sheet("Investments")
    last = write_table(ws, "Investments", INVESTMENT_COLUMNS, investments)
    col = {name: get_column_letter(i) for i, (name, _k, _w) in enumerate(INVESTMENT_COLUMNS, 1)}
    rng = "A2:%s%d" % (get_column_letter(len(INVESTMENT_COLUMNS)), last)
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['$%s2=TRUE' % col["conflict_flag"]], fill=CONFLICT_FILL, stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['OR($%s2="uncertain",$%s2="single_source",$%s2="conflict")'
                 % (col["verification_status"], col["verification_status"],
                    col["verification_status"])],
        fill=UNCERTAIN_FILL))

    # --- Rounds
    ws = wb.create_sheet("Rounds")
    write_table(ws, "Rounds", [
        ("round_id", "txt", 18), ("project_slug", "txt", 26), ("project_name", "txt", 26),
        ("round_date", "date", 13), ("date_precision", "txt", 14), ("round_type", "txt", 14),
        ("round_size_usd", "usd", 16), ("valuation_usd", "usd", 16),
        ("investor_count", "int", 14), ("lead_count", "int", 11),
        ("selected_fund_count", "int", 19), ("selected_funds", "txt", 34),
        ("primary_source_url", "url", 42), ("aggregator_source_url", "url", 42),
        ("all_investors", "txt", 60),
    ], rounds)

    # --- Portfolio Companies
    company_stats = {}
    for inv in investments:
        c = company_stats.setdefault(inv["project_slug"], {"funds": set(), "rounds": set(),
                                                           "first": None, "last": None})
        c["funds"].add(inv["fund_name"])
        c["rounds"].add(inv["round_id"])
        d = inv["round_date"]
        if d:
            c["first"] = d if not c["first"] else min(c["first"], d)
            c["last"] = d if not c["last"] else max(c["last"], d)
    company_rows = []
    for p in projects:
        c = company_stats.get(p["project_slug"], {"funds": set(), "rounds": set(),
                                                  "first": None, "last": None})
        company_rows.append({
            "project_slug": p["project_slug"],
            "project_name": p["project_name"],
            "token_ticker": p["token_ticker"],
            "token_exists": p["token_exists"],
            "selected_fund_count": len(c["funds"]),
            "selected_funds": ", ".join(sorted(c["funds"])),
            "rounds_with_selected_funds": len(c["rounds"]),
            "first_round_date": c["first"],
            "last_round_date": c["last"],
            "project_website": p["project_website"],
            "source_url": p["source_url"],
            "description": p["description"],
        })
    ws = wb.create_sheet("Portfolio Companies")
    write_table(ws, "PortfolioCompanies", [
        ("project_slug", "txt", 26), ("project_name", "txt", 26), ("token_ticker", "txt", 12),
        ("token_exists", "bool", 12), ("selected_fund_count", "int", 19),
        ("selected_funds", "txt", 40), ("rounds_with_selected_funds", "int", 24),
        ("first_round_date", "date", 16), ("last_round_date", "date", 16),
        ("project_website", "url", 34), ("source_url", "url", 42),
        ("description", "txt", 60),
    ], company_rows)

    # --- Coverage
    coverage_rows = []
    for f in dataset["funds"]:
        slug = f["fund_slug"]
        stats = per_fund.get(slug, {"rounds": 0, "companies": set()})
        ctrl = controls.get("funds", {}).get(slug, {})
        cf = ctrl.get("crypto_fundraising") or {}
        cr = ctrl.get("cryptorank") or {}
        off = ctrl.get("official_portfolio") or {}
        rd = ctrl.get("rootdata") or {}
        own = len(stats["companies"])
        cr_n = cr.get("investments")
        diff = (own - cr_n) if isinstance(cr_n, int) else None
        coverage_rows.append({
            "fund_slug": slug,
            "fund_name": f["fund_name"],
            "unique_companies_own_database": own,
            "rounds_own_database": stats["rounds"],
            "crypto_fundraising_fund_page_rows": cf.get("visible_rows"),
            "crypto_fundraising_note": cf.get("note") or cf.get("status") or "",
            "official_portfolio_names": off.get("names_found"),
            "official_portfolio_note": off.get("note") or off.get("status") or "",
            "rootdata_investments": rd.get("investments"),
            "rootdata_note": rd.get("note") or rd.get("status") or "",
            "cryptorank_investments": cr_n,
            "cryptorank_note": cr.get("note") or cr.get("status") or "",
            "difference_own_minus_cryptorank": diff,
            "coverage_statement": COMPLETENESS_STATEMENT,
        })
    ws = wb.create_sheet("Coverage")
    write_table(ws, "Coverage", [
        ("fund_slug", "txt", 20), ("fund_name", "txt", 22),
        ("unique_companies_own_database", "int", 26), ("rounds_own_database", "int", 20),
        ("crypto_fundraising_fund_page_rows", "int", 28), ("crypto_fundraising_note", "txt", 46),
        ("official_portfolio_names", "int", 22), ("official_portfolio_note", "txt", 46),
        ("rootdata_investments", "int", 20), ("rootdata_note", "txt", 46),
        ("cryptorank_investments", "int", 20), ("cryptorank_note", "txt", 46),
        ("difference_own_minus_cryptorank", "int", 28),
        ("coverage_statement", "txt", 60),
    ], coverage_rows)

    # --- Sources
    source_rows = []
    seen = set()

    def add_source(url, kind, rank, context):
        if not url or url in seen:
            return
        seen.add(url)
        source_rows.append({
            "source_url": url, "source_type": kind, "priority_rank": rank,
            "consulted_date": dataset["source_consulted_date"], "context": context,
        })

    for inv in investments:
        add_source(inv["primary_source_url"], "original announcement or press release", 4,
                   "%s — %s" % (inv["project_name"], inv["round_date"] or "date unknown"))
    for inv in investments:
        add_source(inv["aggregator_source_url"], "aggregator crypto-fundraising.info", 5,
                   inv["project_name"])
    for f in dataset["funds"]:
        add_source(f["official_portfolio_url"], "fund portfolio page (control list)", 2,
                   f["fund_name"])
        add_source(f["cryptorank_url"], "aggregator CryptoRank (control list)", 7, f["fund_name"])
        add_source(f["aggregator_fund_url"], "aggregator fund page (shows ten rows)", 5,
                   f["fund_name"])
    ws = wb.create_sheet("Sources")
    write_table(ws, "Sources", [
        ("source_url", "url", 70), ("source_type", "txt", 40), ("priority_rank", "int", 14),
        ("consulted_date", "date", 16), ("context", "txt", 40),
    ], sorted(source_rows, key=lambda r: (r["priority_rank"], r["source_url"])))

    # --- Conflicts
    ws = wb.create_sheet("Conflicts")
    write_table(ws, "Conflicts", [
        ("investment_id", "txt", 18), ("round_id", "txt", 18), ("fund_slug", "txt", 20),
        ("project_name", "txt", 26), ("field", "txt", 18),
        ("value_a", "txt", 18), ("source_a", "url", 44),
        ("value_b", "txt", 18), ("source_b", "txt", 44),
        ("resolution", "txt", 60),
    ], dataset["conflicts"])

    # --- Unknown
    ws = wb.create_sheet("Unknown")
    write_table(ws, "Unknown", [
        ("investment_id", "txt", 18), ("fund_slug", "txt", 20), ("project_name", "txt", 26),
        ("round_date", "date", 13), ("missing_fields", "txt", 40),
        ("attempted", "txt", 60), ("source_url", "url", 44),
    ], dataset["unknowns"])

    # --- Aliases
    ws = wb.create_sheet("Aliases")
    write_table(ws, "Aliases", [
        ("fund_slug", "txt", 20), ("canonical_name", "txt", 22), ("source_slugs", "txt", 28),
        ("aliases", "txt", 44), ("names_seen_in_source", "txt", 44),
        ("alias_evidence", "txt", 70), ("official_portfolio_url", "url", 36),
        ("cryptorank_url", "url", 36),
    ], dataset["aliases"])

    # --- Derived counts as formulas on Funds
    ws = wb["Funds"]
    n = len(fund_rows) + 1
    ws.cell(row=n + 2, column=1, value="Check (formulas over Investments)").font = Font(bold=True)
    ws.cell(row=n + 3, column=1, value="total investment rows")
    ws.cell(row=n + 3, column=2, value="=COUNTA(Investments!A2:A%d)" % (len(investments) + 1))
    ws.cell(row=n + 4, column=1, value="sum of investments_in_database")
    ws.cell(row=n + 4, column=2, value="=SUM(D2:D%d)" % n)
    ws.cell(row=n + 5, column=1, value="total rounds")
    ws.cell(row=n + 5, column=2, value="=COUNTA(Rounds!A2:A%d)" % (len(rounds) + 1))
    ws.cell(row=n + 6, column=1, value="total companies")
    ws.cell(row=n + 6, column=2, value="=COUNTA('Portfolio Companies'!A2:A%d)" % (len(company_rows) + 1))
    ws.cell(row=n + 7, column=1, value="rows with conflict_flag")
    ws.cell(row=n + 7, column=2,
            value="=COUNTIF(Investments!%s2:%s%d,TRUE)" % (col["conflict_flag"],
                                                            col["conflict_flag"], last))

    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    wb.save(xlsx_path)

    if csv_path:
        with open(csv_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow([c[0] for c in INVESTMENT_COLUMNS])
            for inv in investments:
                row = []
                for name, kind, _w in INVESTMENT_COLUMNS:
                    v = inv.get(name)
                    if kind == "bool":
                        row.append("TRUE" if v else "FALSE")
                    elif kind == "usd":
                        row.append("" if v in (None, "", 0) else v)
                    else:
                        row.append("" if v is None else v)
                writer.writerow(row)

    return {
        "investments": len(investments), "rounds": len(rounds),
        "companies": len(company_rows), "sources": len(source_rows),
    }


def main():
    with open(os.path.join(PROCESSED, "dataset.json")) as fh:
        dataset = json.load(fh)
    controls_path = os.path.join(PROCESSED, "coverage-controls.json")
    controls = json.load(open(controls_path)) if os.path.exists(controls_path) else {"funds": {}}

    os.makedirs(OUTPUTS, exist_ok=True)

    stats = build(dataset, controls, XLSX, CSV)
    print("Written: %s" % XLSX)
    print("Written: %s" % CSV)
    print("  investments %d, rounds %d, companies %d, sources %d"
          % (stats["investments"], stats["rounds"], stats["companies"], stats["sources"]))

    os.makedirs(PER_FUND, exist_ok=True)
    print("\nPer fund:")
    for slug, meta in FUNDS.items():
        scoped = scope_to_fund(dataset, slug)
        path = os.path.join(PER_FUND, "vc-investments-%s.xlsx" % slug)
        s = build(scoped, controls, path, fund_slug=slug)
        print("  %-20s %4d investments  %4d rounds  %4d companies  -> %s"
              % (meta["fund_name"], s["investments"], s["rounds"], s["companies"],
                 os.path.basename(path)))


if __name__ == "__main__":
    main()
